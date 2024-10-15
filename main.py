# main.py
import os
import tempfile
import pandas as pd
from fastapi import FastAPI, File, UploadFile, Form
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
import uvicorn
from typing import List
from combine_xls import combine_excel_files, get_column

app = FastAPI()

# Serve static files
app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/", response_class=HTMLResponse)
async def read_root():
    with open("static/index.html", "r") as f:
        content = f.read()
    return content

@app.post("/get_columns")
async def get_columns(file: UploadFile = File(...)):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
        temp_file.write(await file.read())
        temp_file.close()
        df = pd.read_excel(temp_file.name)
        columns = df.columns.tolist()
        os.unlink(temp_file.name)
    return JSONResponse(content={"columns": columns})

@app.post("/combine")
async def combine_files(
    files: List[UploadFile] = File(...),
    columns: List[str] = Form(...),
    case_sensitive: bool = Form(False),
    like_comparison: bool = Form(False),
    debug: bool = Form(False)
):
    # Create temporary files
    temp_files = []
    for file in files:
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        temp_file.write(await file.read())
        temp_file.close()
        temp_files.append(temp_file.name)

    # Create output file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_output:
        # Combine Excel files
        combined_df = None
        for i, (temp_file, column) in enumerate(zip(temp_files, columns)):
            df = pd.read_excel(temp_file)
            if combined_df is None:
                combined_df = df
            else:
                combined_df = pd.merge(
                    combined_df, df,
                    left_on=columns[0], right_on=column,
                    how='inner', suffixes=('', f'_{i}')
                )

        # Save the combined dataframe
        combined_df.to_excel(temp_output.name, index=False)

        # Apply debug highlighting if needed
        if debug:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill

            wb = load_workbook(temp_output.name)
            ws = wb.active

            colors = ['ADD8E6', 'EE82EE', '90EE90', 'FFFACD', 'FFB6C1']  # Light Blue, Violet, Light Green, Light Yellow, Light Pink

            for col in range(1, ws.max_column + 1):
                color = colors[min((col - 1) // len(columns), len(colors) - 1)]
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col).fill = fill

            wb.save(temp_output.name)

        # Return the merged file
        return FileResponse(
            temp_output.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="merged.xlsx"
        )

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)