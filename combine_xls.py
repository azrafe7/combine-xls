import argparse
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def get_column(df, column_spec):
    if column_spec.isdigit():
        return df.columns[int(column_spec)]
    return column_spec

def combine_excel_files(file_a, file_b, column_a, column_b, output_file, case_sensitive=True, like_comparison=False, debug=False):
    # Read Excel files
    df_a = pd.read_excel(file_a)
    df_b = pd.read_excel(file_b)

    # Get actual column names
    col_a = get_column(df_a, column_a)
    col_b = get_column(df_b, column_b)

    # Duplicate columns to be matched to avoid conflicts
    col_a_matched = f"{col_a}_matched"
    col_b_matched = f"{col_b}_matched"
    df_a[col_a_matched] = df_a.loc[:, col_a]
    df_b[col_b_matched] = df_b.loc[:, col_b]

    # Prepare columns for comparison
    if not case_sensitive:
        df_a[col_a_matched] = df_a[col_a].astype(str).str.lower()
        df_b[col_b_matched] = df_b[col_b].astype(str).str.lower()

    # Perform merge based on comparison type
    if not like_comparison:
        merged_df = pd.merge(df_a, df_b, left_on=col_a_matched, right_on=col_b_matched, how='inner', suffixes=('', '_y'))
    else:
        df_a["_join"] = 1
        df_b["_join"] = 1
        merged_df = df_a.merge(df_b, on='_join', suffixes=('', '_y')).drop('_join', axis=1)
        df_b.drop('_join', axis=1, inplace=True)
        merged_df['_match'] = merged_df.apply(lambda x: x[col_a_matched].lower().find(x[col_b_matched].lower()), axis=1).ge(0)
        merged_df = merged_df[merged_df["_match"] == True]
        merged_df.drop("_match", axis=1, inplace=True)

    # Drop duplicated columns to match
    merged_df.drop([col_a_matched, col_b_matched], axis=1, inplace=True)

    # Ensure columns with the same header are included only once
    columns_to_keep = []
    seen_columns = set()
    
    # Add col_a first
    columns_to_keep.append(col_a)
    seen_columns.add(col_a)
    
    # Add the matched column from file B right after col_a
    columns_to_keep.append(col_b)
    seen_columns.add(col_b)
    
    # Add remaining columns from file A
    for col in df_a.columns:
        if col in merged_df.columns and col not in seen_columns:
            columns_to_keep.append(col)
            seen_columns.add(col)
    
    # Add remaining columns from file B
    for col in df_b.columns:
        if col in merged_df.columns and col not in seen_columns:
            columns_to_keep.append(col)
            seen_columns.add(col)

    # Remove any '_y' suffixed columns if their original column exists
    columns_to_keep = [col for col in columns_to_keep if not (col.endswith('_y') and col[:-2] in columns_to_keep)]

    merged_df = merged_df[columns_to_keep]

    # Save the merged dataframe to a new Excel file
    MAX_ROWS = 20
    merged_rows = len(merged_df)
    merged_df = merged_df.head(MAX_ROWS)
    merged_df.to_excel(output_file, index=False)
    
    if debug:
        # Load the workbook and get the active sheet
        wb = load_workbook(output_file)
        ws = wb.active

        # Define fill colors
        fill_match_a = PatternFill(start_color="4DEB69", end_color="4DEB69", fill_type="solid")
        fill_match_b = PatternFill(start_color="FFE316", end_color="FFE316", fill_type="solid")
        fill_both = PatternFill(start_color="DED2EE", end_color="DED2EE", fill_type="solid")
        fill_a = PatternFill(start_color="9EF085", end_color="9EF085", fill_type="solid")
        fill_b = PatternFill(start_color="FAF99A", end_color="FAF99A", fill_type="solid")

        # Create sets of columns from each file
        cols_a = set(df_a.columns)
        cols_b = set(df_b.columns)

        # Apply fill colors
        for col in range(1, ws.max_column + 1):
            col_name = ws.cell(row=1, column=col).value
            if col_name == col_a:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col).fill = fill_match_a
            elif col_name == col_b:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col).fill = fill_match_b
            elif col_name in cols_a and col_name in cols_b:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col).fill = fill_both
            elif col_name in cols_a:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col).fill = fill_a
            elif col_name in cols_b:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col).fill = fill_b

        # Save the workbook
        wb.save(output_file)

    print(f"Combined file saved as {output_file}")
    print(f"Number of rows in file A: {len(df_a)}")
    print(f"Number of rows in file B: {len(df_b)}")
    print(f"Number of rows in merged file: {merged_rows}" + (f" clipped to {MAX_ROWS} for demo purposes" if merged_rows > MAX_ROWS else ""))
    print(f"Columns in merged file: {', '.join(merged_df.columns)}")
    
    return output_file

def main():
    parser = argparse.ArgumentParser(description="Combine two Excel files based on a specified column.")
    parser.add_argument("file_a", help="Path to the first Excel file")
    parser.add_argument("file_b", help="Path to the second Excel file")
    parser.add_argument("--column_a", required=True, help="Column name or index in the first file for comparison")
    parser.add_argument("--column_b", required=True, help="Column name or index in the second file for comparison")
    parser.add_argument("--output", default="merged.xlsx", help="Path to save the combined Excel file (default: merged.xlsx)")
    parser.add_argument("--case-insensitive", action="store_true", help="Perform case-insensitive comparison")
    parser.add_argument("--like", action="store_true", help="Use 'LIKE' comparison instead of exact match")
    parser.add_argument("--debug", action="store_true", help="Highlight cells based on their source file")

    args = parser.parse_args()

    combine_excel_files(args.file_a, args.file_b, args.column_a, args.column_b, 
                        args.output, not args.case_insensitive, args.like, args.debug)

if __name__ == "__main__":
    main()